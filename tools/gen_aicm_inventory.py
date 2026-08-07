#!/usr/bin/env python3
"""Regenerate the CSA AICM requirement inventory from the official workbook.

Not part of the runtime package. Requires openpyxl, which is a tools-only
dependency and is deliberately not in the openaisf runtime dependency set.

The workbook is CSA copyright and is not redistributed with this repository.
Download AICM v1.1 from cloudsecurityalliance.org and pass its path:

    pip install openpyxl
    python tools/gen_aicm_inventory.py "/path/to/AICMv1.1.1-generated_at_2026_07_22.xlsx"

OpenAISF — created by Maarten Loose. Licensed under Apache-2.0.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "spec/crosswalk/inventories/csa_aicm.yaml"
ID_RE = re.compile(r"^[A-Z&]{2,5}-\d{2}$")


def esc(text: str) -> str:
    return '"' + text.replace("\\", "\\\\").replace('"', '\\"') + '"'


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__)
        return 2

    try:
        import openpyxl
    except ImportError:
        print("openpyxl is required: pip install openpyxl", file=sys.stderr)
        return 2

    workbook_path = Path(argv[1])
    if not workbook_path.exists():
        print(f"not found: {workbook_path}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["AICM"]

    # Deliberately reads column C (Control ID) only. Column B (Control Title)
    # and column D (Control Specification) are CSA-authored text and are NOT
    # extracted. See the copyright note written into the output header.
    rows: list[tuple[str, str]] = []
    for r in range(4, ws.max_row + 1):
        cid = ws.cell(r, 3).value
        if cid is None or not ID_RE.match(str(cid).strip()):
            continue
        cid = str(cid).strip()
        domain_code = cid.rsplit("-", 1)[0]
        rows.append((domain_code, cid))

    version = "1.1.1"
    lines = [
        "# CSA AI Controls Matrix (AICM) — REFERENCE-ONLY inventory.",
        "# GENERATED FILE — do not edit by hand.",
        "# Regenerate with: python tools/gen_aicm_inventory.py <workbook.xlsx>",
        "#",
        "# COPYRIGHT POSITION. The AICM is a copyrighted work of the Cloud",
        "# Security Alliance, published on terms that permit free internal and",
        "# non-commercial use but require a CSA licence for commercial use or",
        "# for modified or derivative works. This file therefore contains NO",
        "# CSA-authored text: no control titles, no control specifications, no",
        "# domain names. It carries only control identifiers, which are factual",
        "# designations rather than protectable expression, and the domain code",
        "# derived from each identifier's own prefix.",
        "#",
        "# The enumeration is exhaustive, so it embodies no original selection",
        "# under Feist, and its order follows CSA's own numbering.",
        "#",
        "# This inventory lets OpenAISF account for coverage of the AICM. It is",
        "# not a substitute for it. Obtain the AICM from the Cloud Security",
        "# Alliance to read what any control requires.",
        "#",
        "# ACTION REQUIRED BEFORE COMMERCIAL LAUNCH: obtain a CSA commercial",
        "# licence covering use of AICM identifiers in a commercially offered",
        "# conformance product. See ATTRIBUTIONS.md.",
        "#",
        "# OpenAISF — created by Maarten Loose. Specification licensed CC-BY-4.0.",
        "regime: csa_aicm",
        f"name: CSA AI Controls Matrix v{version}",
        f'version: "{version}"',
        'source: "https://cloudsecurityalliance.org/artifacts/ai-controls-matrix-v1-1"',
        'licence: "CSA proprietary — free internal use, licence required for commercial use"',
        'licence_url: "https://cloudsecurityalliance.org/artifacts/ccm-aicm-licensing-faq"',
        "attribution: >",
        "  Control identifiers refer to the CSA AI Controls Matrix v1.1.1,",
        "  copyright Cloud Security Alliance. No CSA text is reproduced.",
        "  Obtain the AICM from CSA to read it.",
        "regime_kind: requirement",
        "reproduction: reference-only",
        f"declared_total: {len(rows)}",
        "requirements:",
    ]

    last_domain = None
    for domain_code, cid in rows:
        if domain_code != last_domain:
            lines.append(f"  # domain {domain_code}")
            last_domain = domain_code
        lines.append(f"  - ref: {cid}")
        lines.append(
            f"    text_summary: {esc(f'AICM control objective in domain {domain_code}')}"
        )

    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    domains = len({d for d, _ in rows})
    print(f"wrote {OUT.name}: {len(rows)} identifiers across {domains} domains")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
